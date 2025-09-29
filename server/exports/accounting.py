"""
Accounting Export Module
CSV/Excel export with provenance tracking
"""

import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import logging

from ..schemas.invoice import Invoice, ExportRow
from ..audit.logs import log_export

logger = logging.getLogger(__name__)


class AccountingExporter:
    """Exporter for accounting data with provenance tracking"""
    
    def __init__(self):
        self.export_dir = Path("server/exports/output")
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_invoice_to_csv(self, invoice: Invoice, job_id: str) -> str:
        """Export invoice to CSV with provenance"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"invoice_{job_id}_{timestamp}.csv"
        filepath = self.export_dir / filename
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow([
                    'Field Name', 'Field Value', 'Confidence', 'Evidence Page', 
                    'Evidence Bbox', 'Extraction Method', 'Human Reviewed'
                ])
                
                # Export all fields
                rows = self._extract_all_fields(invoice)
                for row in rows:
                    writer.writerow([
                        row.field_name,
                        row.field_value,
                        row.confidence,
                        row.evidence_page,
                        json.dumps(row.evidence_bbox) if row.evidence_bbox else '',
                        row.extraction_method,
                        row.human_reviewed
                    ])
            
            # Log export
            log_export(job_id, 'csv', str(filepath), len(rows))
            
            logger.info(f"✅ Exported invoice to CSV: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Failed to export CSV: {e}")
            raise
    
    def export_invoice_to_excel(self, invoice: Invoice, job_id: str) -> str:
        """Export invoice to Excel with provenance"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"invoice_{job_id}_{timestamp}.xlsx"
        filepath = self.export_dir / filename
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoice Data"
            
            # Style header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write header
            headers = [
                'Field Name', 'Field Value', 'Confidence', 'Evidence Page', 
                'Evidence Bbox', 'Extraction Method', 'Human Reviewed'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Export all fields
            rows = self._extract_all_fields(invoice)
            for row_idx, row in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=row.field_name)
                ws.cell(row=row_idx, column=2, value=row.field_value)
                ws.cell(row=row_idx, column=3, value=row.confidence)
                ws.cell(row=row_idx, column=4, value=row.evidence_page)
                ws.cell(row=row_idx, column=5, value=json.dumps(row.evidence_bbox) if row.evidence_bbox else '')
                ws.cell(row=row_idx, column=6, value=row.extraction_method)
                ws.cell(row=row_idx, column=7, value=row.human_reviewed)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            # Log export
            log_export(job_id, 'excel', str(filepath), len(rows))
            
            logger.info(f"✅ Exported invoice to Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Failed to export Excel: {e}")
            raise
    
    def export_multiple_invoices(self, invoices: List[Invoice], job_ids: List[str], 
                               format: str = 'csv') -> str:
        """Export multiple invoices to single file"""
        if format == 'csv':
            return self._export_multiple_to_csv(invoices, job_ids)
        elif format == 'excel':
            return self._export_multiple_to_excel(invoices, job_ids)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _export_multiple_to_csv(self, invoices: List[Invoice], job_ids: List[str]) -> str:
        """Export multiple invoices to CSV"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"invoices_batch_{timestamp}.csv"
        filepath = self.export_dir / filename
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow([
                    'Job ID', 'Invoice Number', 'Vendor', 'Invoice Date', 'Grand Total',
                    'Field Name', 'Field Value', 'Confidence', 'Evidence Page', 
                    'Evidence Bbox', 'Extraction Method', 'Human Reviewed'
                ])
                
                # Export all invoices
                total_rows = 0
                for invoice, job_id in zip(invoices, job_ids):
                    rows = self._extract_all_fields(invoice)
                    for row in rows:
                        writer.writerow([
                            job_id,
                            invoice.invoice_number.value,
                            invoice.vendor.name.value,
                            invoice.invoice_date.value,
                            invoice.amounts.grand_total.value,
                            row.field_name,
                            row.field_value,
                            row.confidence,
                            row.evidence_page,
                            json.dumps(row.evidence_bbox) if row.evidence_bbox else '',
                            row.extraction_method,
                            row.human_reviewed
                        ])
                        total_rows += 1
            
            # Log export
            log_export("batch", 'csv', str(filepath), total_rows)
            
            logger.info(f"✅ Exported {len(invoices)} invoices to CSV: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Failed to export batch CSV: {e}")
            raise
    
    def _export_multiple_to_excel(self, invoices: List[Invoice], job_ids: List[str]) -> str:
        """Export multiple invoices to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"invoices_batch_{timestamp}.xlsx"
        filepath = self.export_dir / filename
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices Data"
            
            # Style header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write header
            headers = [
                'Job ID', 'Invoice Number', 'Vendor', 'Invoice Date', 'Grand Total',
                'Field Name', 'Field Value', 'Confidence', 'Evidence Page', 
                'Evidence Bbox', 'Extraction Method', 'Human Reviewed'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Export all invoices
            row_idx = 2
            total_rows = 0
            for invoice, job_id in zip(invoices, job_ids):
                rows = self._extract_all_fields(invoice)
                for row in rows:
                    ws.cell(row=row_idx, column=1, value=job_id)
                    ws.cell(row=row_idx, column=2, value=invoice.invoice_number.value)
                    ws.cell(row=row_idx, column=3, value=invoice.vendor.name.value)
                    ws.cell(row=row_idx, column=4, value=invoice.invoice_date.value)
                    ws.cell(row=row_idx, column=5, value=invoice.amounts.grand_total.value)
                    ws.cell(row=row_idx, column=6, value=row.field_name)
                    ws.cell(row=row_idx, column=7, value=row.field_value)
                    ws.cell(row=row_idx, column=8, value=row.confidence)
                    ws.cell(row=row_idx, column=9, value=row.evidence_page)
                    ws.cell(row=row_idx, column=10, value=json.dumps(row.evidence_bbox) if row.evidence_bbox else '')
                    ws.cell(row=row_idx, column=11, value=row.extraction_method)
                    ws.cell(row=row_idx, column=12, value=row.human_reviewed)
                    row_idx += 1
                    total_rows += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            # Log export
            log_export("batch", 'excel', str(filepath), total_rows)
            
            logger.info(f"✅ Exported {len(invoices)} invoices to Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Failed to export batch Excel: {e}")
            raise
    
    def _extract_all_fields(self, invoice: Invoice) -> List[ExportRow]:
        """Extract all fields from invoice with provenance"""
        rows = []
        
        # Invoice identifiers
        rows.append(ExportRow(
            field_name='invoice_number',
            field_value=invoice.invoice_number.value,
            confidence=invoice.invoice_number.confidence,
            evidence_page=invoice.invoice_number.evidence[0].page if invoice.invoice_number.evidence else None,
            evidence_bbox=invoice.invoice_number.evidence[0].bbox if invoice.invoice_number.evidence else None,
            extraction_method=invoice.extraction_method,
            human_reviewed=invoice.human_reviewed
        ))
        
        rows.append(ExportRow(
            field_name='invoice_date',
            field_value=invoice.invoice_date.value,
            confidence=invoice.invoice_date.confidence,
            evidence_page=invoice.invoice_date.evidence[0].page if invoice.invoice_date.evidence else None,
            evidence_bbox=invoice.invoice_date.evidence[0].bbox if invoice.invoice_date.evidence else None,
            extraction_method=invoice.extraction_method,
            human_reviewed=invoice.human_reviewed
        ))
        
        if invoice.due_date:
            rows.append(ExportRow(
                field_name='due_date',
                field_value=invoice.due_date.value,
                confidence=invoice.due_date.confidence,
                evidence_page=invoice.due_date.evidence[0].page if invoice.due_date.evidence else None,
                evidence_bbox=invoice.due_date.evidence[0].bbox if invoice.due_date.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        # Vendor information
        rows.append(ExportRow(
            field_name='vendor_name',
            field_value=invoice.vendor.name.value,
            confidence=invoice.vendor.name.confidence,
            evidence_page=invoice.vendor.name.evidence[0].page if invoice.vendor.name.evidence else None,
            evidence_bbox=invoice.vendor.name.evidence[0].bbox if invoice.vendor.name.evidence else None,
            extraction_method=invoice.extraction_method,
            human_reviewed=invoice.human_reviewed
        ))
        
        if invoice.vendor.address:
            rows.append(ExportRow(
                field_name='vendor_address',
                field_value=invoice.vendor.address.value,
                confidence=invoice.vendor.address.confidence,
                evidence_page=invoice.vendor.address.evidence[0].page if invoice.vendor.address.evidence else None,
                evidence_bbox=invoice.vendor.address.evidence[0].bbox if invoice.vendor.address.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.vendor.tax_id:
            rows.append(ExportRow(
                field_name='vendor_tax_id',
                field_value=invoice.vendor.tax_id.value,
                confidence=invoice.vendor.tax_id.confidence,
                evidence_page=invoice.vendor.tax_id.evidence[0].page if invoice.vendor.tax_id.evidence else None,
                evidence_bbox=invoice.vendor.tax_id.evidence[0].bbox if invoice.vendor.tax_id.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        # Financial amounts
        rows.append(ExportRow(
            field_name='grand_total',
            field_value=invoice.amounts.grand_total.value,
            confidence=invoice.amounts.grand_total.confidence,
            evidence_page=invoice.amounts.grand_total.evidence[0].page if invoice.amounts.grand_total.evidence else None,
            evidence_bbox=invoice.amounts.grand_total.evidence[0].bbox if invoice.amounts.grand_total.evidence else None,
            extraction_method=invoice.extraction_method,
            human_reviewed=invoice.human_reviewed
        ))
        
        rows.append(ExportRow(
            field_name='currency',
            field_value=invoice.amounts.currency.value,
            confidence=invoice.amounts.currency.confidence,
            evidence_page=invoice.amounts.currency.evidence[0].page if invoice.amounts.currency.evidence else None,
            evidence_bbox=invoice.amounts.currency.evidence[0].bbox if invoice.amounts.currency.evidence else None,
            extraction_method=invoice.extraction_method,
            human_reviewed=invoice.human_reviewed
        ))
        
        if invoice.amounts.subtotal:
            rows.append(ExportRow(
                field_name='subtotal',
                field_value=invoice.amounts.subtotal.value,
                confidence=invoice.amounts.subtotal.confidence,
                evidence_page=invoice.amounts.subtotal.evidence[0].page if invoice.amounts.subtotal.evidence else None,
                evidence_bbox=invoice.amounts.subtotal.evidence[0].bbox if invoice.amounts.subtotal.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.amounts.tax_amount:
            rows.append(ExportRow(
                field_name='tax_amount',
                field_value=invoice.amounts.tax_amount.value,
                confidence=invoice.amounts.tax_amount.confidence,
                evidence_page=invoice.amounts.tax_amount.evidence[0].page if invoice.amounts.tax_amount.evidence else None,
                evidence_bbox=invoice.amounts.tax_amount.evidence[0].bbox if invoice.amounts.tax_amount.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.amounts.tax_rate:
            rows.append(ExportRow(
                field_name='tax_rate',
                field_value=invoice.amounts.tax_rate.value,
                confidence=invoice.amounts.tax_rate.confidence,
                evidence_page=invoice.amounts.tax_rate.evidence[0].page if invoice.amounts.tax_rate.evidence else None,
                evidence_bbox=invoice.amounts.tax_rate.evidence[0].bbox if invoice.amounts.tax_rate.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.amounts.discount:
            rows.append(ExportRow(
                field_name='discount',
                field_value=invoice.amounts.discount.value,
                confidence=invoice.amounts.discount.confidence,
                evidence_page=invoice.amounts.discount.evidence[0].page if invoice.amounts.discount.evidence else None,
                evidence_bbox=invoice.amounts.discount.evidence[0].bbox if invoice.amounts.discount.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.amounts.shipping:
            rows.append(ExportRow(
                field_name='shipping',
                field_value=invoice.amounts.shipping.value,
                confidence=invoice.amounts.shipping.confidence,
                evidence_page=invoice.amounts.shipping.evidence[0].page if invoice.amounts.shipping.evidence else None,
                evidence_bbox=invoice.amounts.shipping.evidence[0].bbox if invoice.amounts.shipping.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        # Line items
        for i, line_item in enumerate(invoice.line_items):
            rows.append(ExportRow(
                field_name=f'line_item_{i+1}_description',
                field_value=line_item.description.value,
                confidence=line_item.description.confidence,
                evidence_page=line_item.description.evidence[0].page if line_item.description.evidence else None,
                evidence_bbox=line_item.description.evidence[0].bbox if line_item.description.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
            
            if line_item.quantity:
                rows.append(ExportRow(
                    field_name=f'line_item_{i+1}_quantity',
                    field_value=line_item.quantity.value,
                    confidence=line_item.quantity.confidence,
                    evidence_page=line_item.quantity.evidence[0].page if line_item.quantity.evidence else None,
                    evidence_bbox=line_item.quantity.evidence[0].bbox if line_item.quantity.evidence else None,
                    extraction_method=invoice.extraction_method,
                    human_reviewed=invoice.human_reviewed
                ))
            
            if line_item.unit_price:
                rows.append(ExportRow(
                    field_name=f'line_item_{i+1}_unit_price',
                    field_value=line_item.unit_price.value,
                    confidence=line_item.unit_price.confidence,
                    evidence_page=line_item.unit_price.evidence[0].page if line_item.unit_price.evidence else None,
                    evidence_bbox=line_item.unit_price.evidence[0].bbox if line_item.unit_price.evidence else None,
                    extraction_method=invoice.extraction_method,
                    human_reviewed=invoice.human_reviewed
                ))
            
            if line_item.total:
                rows.append(ExportRow(
                    field_name=f'line_item_{i+1}_total',
                    field_value=line_item.total.value,
                    confidence=line_item.total.confidence,
                    evidence_page=line_item.total.evidence[0].page if line_item.total.evidence else None,
                    evidence_bbox=line_item.total.evidence[0].bbox if line_item.total.evidence else None,
                    extraction_method=invoice.extraction_method,
                    human_reviewed=invoice.human_reviewed
                ))
            
            if line_item.category:
                rows.append(ExportRow(
                    field_name=f'line_item_{i+1}_category',
                    field_value=line_item.category,
                    confidence=line_item.category_confidence or 0.0,
                    evidence_page=None,
                    evidence_bbox=None,
                    extraction_method='ml_classification',
                    human_reviewed=invoice.human_reviewed
                ))
        
        # Additional fields
        if invoice.notes:
            rows.append(ExportRow(
                field_name='notes',
                field_value=invoice.notes.value,
                confidence=invoice.notes.confidence,
                evidence_page=invoice.notes.evidence[0].page if invoice.notes.evidence else None,
                evidence_bbox=invoice.notes.evidence[0].bbox if invoice.notes.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.payment_terms:
            rows.append(ExportRow(
                field_name='payment_terms',
                field_value=invoice.payment_terms.value,
                confidence=invoice.payment_terms.confidence,
                evidence_page=invoice.payment_terms.evidence[0].page if invoice.payment_terms.evidence else None,
                evidence_bbox=invoice.payment_terms.evidence[0].bbox if invoice.payment_terms.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        if invoice.po_number:
            rows.append(ExportRow(
                field_name='po_number',
                field_value=invoice.po_number.value,
                confidence=invoice.po_number.confidence,
                evidence_page=invoice.po_number.evidence[0].page if invoice.po_number.evidence else None,
                evidence_bbox=invoice.po_number.evidence[0].bbox if invoice.po_number.evidence else None,
                extraction_method=invoice.extraction_method,
                human_reviewed=invoice.human_reviewed
            ))
        
        return rows


# Global exporter instance
exporter = AccountingExporter()


def export_invoice_to_csv(invoice: Invoice, job_id: str) -> str:
    """Export invoice to CSV"""
    return exporter.export_invoice_to_csv(invoice, job_id)


def export_invoice_to_excel(invoice: Invoice, job_id: str) -> str:
    """Export invoice to Excel"""
    return exporter.export_invoice_to_excel(invoice, job_id)


def export_multiple_invoices(invoices: List[Invoice], job_ids: List[str], 
                           format: str = 'csv') -> str:
    """Export multiple invoices"""
    return exporter.export_multiple_invoices(invoices, job_ids, format)







